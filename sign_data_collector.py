#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook, load_workbook

EXCEL_PATH = Path(r"C:\Users\ideapad15s\Desktop\art_pi2_wifi_project\data_raw.xlsx")
START_ROW = 10
PASTE_COUNT_PER_ACTION = 60
PAIR_COUNT_PER_ACTION = 30
ALLOWED_HANDS = {"both", "right", "left"}
SENSOR_VALUE_COUNT = 69
TOTAL_FIELDS_PER_DATA = 2 + SENSOR_VALUE_COUNT
END_MARKER = "END"
TIMESTAMP_GAP_THRESHOLD = 200  # 时间戳间隔阈值（毫秒），超过此值认为是不同动作


def clean_text(text: str) -> str:
    text = text.replace("，", ",")
    text = re.sub(r"<[^>]*>", "", text)
    text = re.sub(r"\s+", "", text)
    return text.strip()


def parse_data_records(block: str) -> Tuple[List[Tuple[int, str, List[int]]], str]:
    pattern = r'\[DATA\]([^\[\n]+)'
    matches = re.findall(pattern, block)

    if not matches:
        return [], "未解析到任何 [DATA] 记录。"

    records: List[Tuple[int, str, List[int]]] = []
    skipped_count = 0

    for match in matches:
        segment = clean_text(match)
        if not segment:
            skipped_count += 1
            continue

        fields = [x for x in segment.split(",") if x != ""]
        if len(fields) != TOTAL_FIELDS_PER_DATA:
            skipped_count += 1
            continue

        try:
            timestamp_ms = int(fields[0])
        except ValueError:
            skipped_count += 1
            continue

        hand = fields[1].lower()
        if hand not in ALLOWED_HANDS:
            skipped_count += 1
            continue

        sensor_values: List[int] = []
        valid_row = True
        for value in fields[2:]:
            try:
                sensor_values.append(int(value))
            except ValueError:
                valid_row = False
                break

        if not valid_row:
            skipped_count += 1
            continue

        records.append((timestamp_ms, hand, sensor_values))

    if not records:
        return [], f"未解析到有效 [DATA] 记录（跳过 {skipped_count} 条格式错误数据）。"

    if skipped_count > 0:
        print(f"提示：本次粘贴自动跳过 {skipped_count} 条格式错误的 [DATA] 行。")

    return records, ""


def load_or_create_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    wb.save(path)
    return wb, ws


def find_next_row(ws) -> int:
    last_used = 0
    for r in range(ws.max_row, START_ROW - 1, -1):
        row_has_value = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(2, 79))
        if row_has_value:
            last_used = r
            break

    if last_used == 0:
        return START_ROW
    return last_used + 1


def next_gt_id(ws) -> int:
    max_id = 0
    for r in range(1, ws.max_row + 1):
        value = ws.cell(row=r, column=6).value
        if isinstance(value, str):
            m = re.fullmatch(r"gt_(\d+)", value.strip())
            if m:
                max_id = max(max_id, int(m.group(1)))
    return max_id + 1


def input_non_empty(prompt: str) -> str:
    while True:
        value = input(prompt).strip()
        if value:
            return value
        print("输入不能为空，请重新输入。")


def input_action_hand(prompt: str) -> str:
    while True:
        value = input(prompt).strip().lower()
        if value in ALLOWED_HANDS:
            return value
        print("使用手输入错误，只能是 both / right / left。")


def select_mode() -> str:
    print("\n请选择采集模式：")
    print("1. 手动采集")
    print("2. 自动采集")
    while True:
        choice = input("请输入选项（1或2）: ").strip()
        if choice == "1":
            return "manual"
        elif choice == "2":
            return "auto"
        print("输入错误，请输入 1 或 2。")


def read_block(expected_hand: str) -> str:
    print(f"请粘贴期望手为 {expected_hand} 的数据块，粘贴后直接回车：")
    line = input()
    return line


def parse_expected_hand_block(expected_hand: str) -> Tuple[List[Tuple[int, List[int]]], str]:
    block = read_block(expected_hand)
    records, err = parse_data_records(block)
    if err:
        return [], err

    if any(hand != expected_hand for _, hand, _ in records):
        return [], f"数据中的手字段不全是 {expected_hand}。"

    return [(timestamp_ms, sensor_values) for timestamp_ms, _, sensor_values in records], ""


def merge_left_right_for_write(left_records: List[Tuple[int, List[int]]], right_records: List[Tuple[int, List[int]]]):
    common_count = min(len(left_records), len(right_records))
    merged = []
    for t in range(common_count):
        left_timestamp, left_values = left_records[t]
        right_timestamp, right_values = right_records[t]
        merged.append((t, left_timestamp, "left", left_values))
        merged.append((t, right_timestamp, "right", right_values))
    return merged, common_count


def speed_by_pair_index(pair_index: int) -> str:
    if pair_index <= 10:
        return "slow"
    if pair_index <= 20:
        return "medium"
    return "high"


def read_file_data(file_path: str, expected_hand: str) -> Tuple[List[Tuple[int, List[int]]], str]:
    # 去除路径两端的引号（单引号或双引号）
    file_path = file_path.strip().strip('"').strip("'")

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        return [], f"文件不存在: {file_path}"
    except Exception as e:
        return [], f"读取文件失败: {e}"

    records, err = parse_data_records(content)
    if err:
        return [], err

    if any(hand != expected_hand for _, hand, _ in records):
        return [], f"文件中的手字段不全是 {expected_hand}。"

    return [(timestamp_ms, sensor_values) for timestamp_ms, _, sensor_values in records], ""


def split_by_timestamp_gap(records: List[Tuple[int, List[int]]]) -> List[List[Tuple[int, List[int]]]]:
    if not records:
        return []

    groups = []
    current_group = [records[0]]

    for i in range(1, len(records)):
        prev_timestamp = records[i - 1][0]
        curr_timestamp = records[i][0]

        if curr_timestamp - prev_timestamp > TIMESTAMP_GAP_THRESHOLD:
            groups.append(current_group)
            current_group = [records[i]]
        else:
            current_group.append(records[i])

    if current_group:
        groups.append(current_group)

    return groups


def find_best_alignment(left_data: List[Tuple[int, List[int]]], right_data: List[Tuple[int, List[int]]]):
    """
    找到左右手数据的最佳对齐方式，基于时间戳差值的稳定性

    参数：
        left_data: [(timestamp, sensor_values), ...]
        right_data: [(timestamp, sensor_values), ...]

    返回：
        (left_start, right_start, match_count, std_dev, mean_diff)
        如果无法找到合适的对齐方式，返回 None
    """
    best_std = float('inf')
    best_result = None

    # 尝试不同的起始位置，最多跳过10条数据
    max_skip = min(10, len(left_data) - 1, len(right_data) - 1)

    for left_start in range(max_skip + 1):
        for right_start in range(max_skip + 1):
            # 计算可以匹配的数据条数
            match_count = min(len(left_data) - left_start, len(right_data) - right_start)

            # 至少要有5条数据才有意义
            if match_count < 5:
                continue

            # 计算时间戳差值序列
            time_diffs = []
            for i in range(match_count):
                left_ts = left_data[left_start + i][0]
                right_ts = right_data[right_start + i][0]
                time_diffs.append(left_ts - right_ts)

            # 计算时间戳差值的平均值和标准差
            mean_diff = sum(time_diffs) / len(time_diffs)
            variance = sum((d - mean_diff) ** 2 for d in time_diffs) / len(time_diffs)
            std_dev = variance ** 0.5

            # 更新最佳结果（标准差越小越好）
            if std_dev < best_std:
                best_std = std_dev
                best_result = (left_start, right_start, match_count, std_dev, mean_diff)

    return best_result


def align_action_groups(left_groups, right_groups):
    """
    对齐左右手的动作实例序列，处理跨动作实例的错位问题

    返回：
        [(left_group, right_group), ...]
        已对齐的动作实例对列表
    """
    paired_groups = []
    left_idx = 0
    right_idx = 0

    print("\n开始动作实例序列对齐...")

    while left_idx < len(left_groups) and right_idx < len(right_groups):
        left_group = left_groups[left_idx]
        right_group = right_groups[right_idx]

        left_count = len(left_group)
        right_count = len(right_group)
        count_diff = abs(left_count - right_count)

        # 如果数据条数差异很大（>10），检查是否需要跳过某个动作实例
        if count_diff > 10:
            should_skip_right = False
            should_skip_left = False

            # 检查是否应该跳过当前right_group
            if right_idx + 1 < len(right_groups):
                next_right_group = right_groups[right_idx + 1]
                next_right_count = len(next_right_group)
                next_count_diff = abs(left_count - next_right_count)

                # 如果下一个right_group与当前left_group更匹配，跳过当前right_group
                if next_count_diff < count_diff and next_count_diff <= 5:
                    should_skip_right = True

            # 检查是否应该跳过当前left_group
            if not should_skip_right and left_idx + 1 < len(left_groups):
                next_left_group = left_groups[left_idx + 1]
                next_left_count = len(next_left_group)
                next_count_diff = abs(next_left_count - right_count)

                # 如果下一个left_group与当前right_group更匹配，跳过当前left_group
                if next_count_diff < count_diff and next_count_diff <= 5:
                    should_skip_left = True

            if should_skip_right:
                print(f"  跳过 right 动作实例 #{right_idx + 1}（数据条数={right_count}，与 left #{left_idx + 1} 不匹配）")
                right_idx += 1
                continue

            if should_skip_left:
                print(f"  跳过 left 动作实例 #{left_idx + 1}（数据条数={left_count}，与 right #{right_idx + 1} 不匹配）")
                left_idx += 1
                continue

        # 配对成功
        paired_groups.append((left_group, right_group))
        left_idx += 1
        right_idx += 1

    print(f"动作实例序列对齐完成，成功配对 {len(paired_groups)} 个动作实例。")

    if left_idx < len(left_groups):
        print(f"  剩余 {len(left_groups) - left_idx} 个 left 动作实例未配对。")

    if right_idx < len(right_groups):
        print(f"  剩余 {len(right_groups) - right_idx} 个 right 动作实例未配对。")

    return paired_groups


def write_records(
    ws,
    start_row: int,
    collector: str,
    speed: str,
    meaning: str,
    action_hand: str,
    gt_label: str,
    merged_records,
):
    row = start_row
    for frame_id, timestamp_ms, data_hand, sensor_values in merged_records:
        ws.cell(row=row, column=2, value=collector)
        ws.cell(row=row, column=3, value=speed)
        ws.cell(row=row, column=4, value=meaning)
        ws.cell(row=row, column=5, value=action_hand)
        ws.cell(row=row, column=6, value=gt_label)
        ws.cell(row=row, column=7, value=timestamp_ms)
        ws.cell(row=row, column=8, value=frame_id)
        ws.cell(row=row, column=9, value=data_hand)

        for offset, sensor in enumerate(sensor_values):
            ws.cell(row=row, column=10 + offset, value=sensor)

        row += 1
    return row


def run_manual_collect():
    wb, ws = load_or_create_workbook(EXCEL_PATH)
    current_row = find_next_row(ws)
    gt_counter = next_gt_id(ws)

    print(f"Excel文件: {EXCEL_PATH}")
    print(f"当前写入起始行: {current_row}")
    print("提示：按 Ctrl+C 可以安全退出程序\n")

    collector = input_non_empty("请输入采集者名字: ")

    while True:
        meaning = input_non_empty("\n请输入动作中文含义（输入 q 退出）: ")
        if meaning.lower() == "q":
            break

        action_hand = input_action_hand("请输入动作使用手（both/right/left）: ")

        print(f"开始采集：固定 {PASTE_COUNT_PER_ACTION} 次粘贴 = {PAIR_COUNT_PER_ACTION} 轮配对。")
        print("每轮必须先粘 left，再粘 right，完成后按 times 从小到大写入：left 在前，right 在后。")

        pair_index = 1
        while pair_index <= PAIR_COUNT_PER_ACTION:
            gt_label = f"gt_{gt_counter:03d}"
            print(f"\n第 {pair_index}/{PAIR_COUNT_PER_ACTION} 轮配对（动作标志: {gt_label}）：")


            left_records, err = parse_expected_hand_block("left")
            if err:
                print(f"left 数据块错误：{err} 本轮无效，请重新开始本轮。")
                continue

            right_records, err = parse_expected_hand_block("right")
            if err:
                print(f"right 数据块错误：{err} 本轮无效，请重新开始本轮。")
                continue

            merged_records, common_count = merge_left_right_for_write(left_records, right_records)
            if common_count == 0:
                print("left/right 无可配对数据（交集为0），本轮无效，请重新开始本轮。")
                continue

            if len(left_records) != len(right_records):
                print(
                    f"警告：left 条数={len(left_records)}，right 条数={len(right_records)}，"
                    f"仅写入交集 {common_count} 组 frame。"
                )

            speed = speed_by_pair_index(pair_index)
            current_row = write_records(
                ws=ws,
                start_row=current_row,
                collector=collector,
                speed=speed,
                meaning=meaning,
                action_hand=action_hand,
                gt_label=gt_label,
                merged_records=merged_records,
            )
            wb.save(EXCEL_PATH)
            print(f"本轮成功写入 {len(merged_records)} 行（{common_count} 组 times）。")

            gt_counter += 1
            pair_index += 1

        wb.save(EXCEL_PATH)
        print(f"动作 {meaning} 采集完成（{PAIR_COUNT_PER_ACTION} 轮配对，共{PASTE_COUNT_PER_ACTION}次粘贴）。")

        go_on = input("是否继续录入下一个动作？(y/n): ").strip().lower()
        if go_on != "y":
            break

    wb.save(EXCEL_PATH)
    print("脚本结束，数据已保存。")


def run_auto_collect():
    wb, ws = load_or_create_workbook(EXCEL_PATH)
    current_row = find_next_row(ws)
    gt_counter = next_gt_id(ws)

    print(f"Excel文件: {EXCEL_PATH}")
    print(f"当前写入起始行: {current_row}")
    print("提示：按 Ctrl+C 可以安全退出程序\n")

    collector = input_non_empty("请输入采集者名字: ")

    while True:
        meaning = input_non_empty("\n请输入动作中文含义（输入 q 退出）: ")
        if meaning.lower() == "q":
            break

        action_hand = input_action_hand("请输入动作使用手（both/right/left）: ")

        left_file_path = input_non_empty("请输入左手文档路径: ")
        right_file_path = input_non_empty("请输入右手文档路径: ")

        print("\n正在读取文件...")
        left_records, err = read_file_data(left_file_path, "left")
        if err:
            print(f"左手文件读取失败：{err}")
            continue

        right_records, err = read_file_data(right_file_path, "right")
        if err:
            print(f"右手文件读取失败：{err}")
            continue

        print(f"左手文件读取成功，共 {len(left_records)} 条数据。")
        print(f"右手文件读取成功，共 {len(right_records)} 条数据。")

        print("\n正在根据时间戳间隔分割动作实例...")
        left_groups = split_by_timestamp_gap(left_records)
        right_groups = split_by_timestamp_gap(right_records)

        print(f"左手文件分割为 {len(left_groups)} 个动作实例。")
        print(f"右手文件分割为 {len(right_groups)} 个动作实例。")

        # 对齐左右手动作实例序列
        paired_groups = align_action_groups(left_groups, right_groups)

        action_count = len(paired_groups)
        if action_count == 0:
            print("错误：没有可配对的动作实例。")
            continue

        print(f"\n开始写入 {action_count} 个动作实例...")

        for pair_index in range(1, action_count + 1):
            gt_label = f"gt_{gt_counter:03d}"
            left_group, right_group = paired_groups[pair_index - 1]

            # 如果数据条数差异 > 3，启用纠错机制
            original_left_count = len(left_group)
            original_right_count = len(right_group)

            if abs(original_left_count - original_right_count) > 3:
                print(f"\n第 {pair_index} 个动作实例数据条数差异较大（left={original_left_count}, right={original_right_count}），启用纠错机制...")

                alignment = find_best_alignment(left_group, right_group)

                if alignment:
                    left_start, right_start, match_count, std_dev, mean_diff = alignment

                    # 截取对齐后的数据
                    left_group = left_group[left_start:left_start + match_count]
                    right_group = right_group[right_start:right_start + match_count]

                    print(f"  纠错完成：跳过左手前 {left_start} 条，跳过右手前 {right_start} 条")
                    print(f"  时间戳差值：平均 {mean_diff:.1f}ms，标准差 {std_dev:.1f}ms")

                    if left_start > 0:
                        skipped_timestamps = [left_groups[pair_index - 1][i][0] for i in range(left_start)]
                        print(f"  左手跳过的时间戳: {skipped_timestamps}")

                    if right_start > 0:
                        skipped_timestamps = [right_groups[pair_index - 1][i][0] for i in range(right_start)]
                        print(f"  右手跳过的时间戳: {skipped_timestamps}")
                else:
                    print(f"  警告：无法找到合适的对齐方式，将使用原始数据取交集。")

            merged_records, common_count = merge_left_right_for_write(left_group, right_group)
            if common_count == 0:
                print(f"第 {pair_index} 个动作实例无可配对数据，跳过。")
                continue

            if len(left_group) != len(right_group):
                print(
                    f"第 {pair_index} 个动作实例：left 条数={len(left_group)}，right 条数={len(right_group)}，"
                    f"仅写入交集 {common_count} 组 frame。"
                )
            else:
                print(f"第 {pair_index} 个动作实例：left 条数={len(left_group)}，right 条数={len(right_group)}，写入 {common_count} 组 frame。")

            speed = speed_by_pair_index(pair_index)
            current_row = write_records(
                ws=ws,
                start_row=current_row,
                collector=collector,
                speed=speed,
                meaning=meaning,
                action_hand=action_hand,
                gt_label=gt_label,
                merged_records=merged_records,
            )

            gt_counter += 1

        wb.save(EXCEL_PATH)
        print(f"\n动作 {meaning} 采集完成，共写入 {action_count} 个动作实例。")

        go_on = input("是否继续录入下一个动作？(y/n): ").strip().lower()
        if go_on != "y":
            break

    wb.save(EXCEL_PATH)
    print("脚本结束，数据已保存。")


def run():
    mode = select_mode()
    if mode == "manual":
        run_manual_collect()
    else:
        run_auto_collect()


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\n\n检测到 Ctrl+C 中断信号。")
        confirm = input("确定要退出吗？已采集的数据已自动保存。(y/n): ").strip().lower()
        if confirm == "y":
            print("程序已安全退出。")
        else:
            print("继续运行...")
            run()
