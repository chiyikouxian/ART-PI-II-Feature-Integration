import openpyxl
from openpyxl.utils import get_column_letter
import sys
import re

# Excel文件路径
EXCEL_PATH = r"C:\Users\ideapad15s\Downloads\data_raw.xlsx"

def load_excel():
    """加载Excel文件，返回工作簿和工作表"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        return wb, ws
    except FileNotFoundError:
        print(f"错误：找不到文件 {EXCEL_PATH}")
        sys.exit(1)
    except Exception as e:
        print(f"加载Excel文件出错：{e}")
        sys.exit(1)

def get_header_columns(ws):
    """获取表头列数"""
    # 第2行（索引1）是实际表头
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    return len(header_row)

def get_target_start_row(ws):
    """获取数据写入的起始行：固定从第10行开始，有数据则往下找第一个空行"""
    base_start_row = 10
    # 检查B列（第2列）是否为空来判断行是否可用
    for row in range(base_start_row, ws.max_row + 2):
        if ws.cell(row=row, column=2).value is None:
            return row
    return base_start_row

def get_last_sample_counter(ws):
    """【新增】自动读取最后一行的动作标志，返回下一个计数器值"""
    # 从最后一行往上找，直到找到有数据的行
    for row in range(ws.max_row, 1, -1):
        # 检查E列（第5列）是否有动作标志（sample_id）
        sample_id_cell = ws.cell(row=row, column=5).value
        if sample_id_cell and isinstance(sample_id_cell, str):
            # 用正则提取数字部分，比如 gt_003 -> 3
            match = re.search(r'gt_(\d+)', sample_id_cell, re.IGNORECASE)
            if match:
                last_num = int(match.group(1))
                return last_num + 1
    # 如果没找到任何历史数据，从1开始
    return 1

def parse_input_data(input_str):
    """解析输入数据，不再严格校验列数"""
    data_lines = []
    
    # 处理多条数据标记[DATA]
    if "[DATA]" in input_str:
        raw_lines = input_str.split("[DATA]")
        raw_lines = [line.strip() for line in raw_lines if line.strip()]
    else:
        raw_lines = [input_str.strip()]
    
    for line in raw_lines:
        # 移除换行符，按逗号分割
        parts = [p.strip() for p in line.replace("\n", "").split(",") if p.strip()]
        
        # 只要数据列数大于60（合理范围）就接受
        if len(parts) < 60:
            print(f"警告：数据列数过少（仅{len(parts)}列），跳过该行。")
            continue
        
        data_lines.append(parts)
    
    return data_lines

def main():
    print("=== 数据集采集工具 ===")
    
    # 加载Excel
    wb, ws = load_excel()
    total_cols = get_header_columns(ws)
    print(f"Excel表头列数：{total_cols}")
    
    # 【修改后】自动获取起始计数器
    sample_counter = get_last_sample_counter(ws)
    if sample_counter == 1:
        print("未检测到历史数据，将从 gt_001 开始。")
    else:
        print(f"检测到历史数据，将从 gt_{sample_counter:03d} 开始。")
    
    # 一次性输入两个参数
    while True:
        init_input = input("\n请一次性输入参数（格式：采集者,中文标签）：\n>>> ").strip()
        parts = [p.strip() for p in init_input.split(",") if p.strip()]

        if len(parts) >= 2:
            worker = parts[0]
            label = parts[1]
            break
        else:
            print("❌ 输入格式错误！请确保包含两个部分，用英文逗号分隔。")
            print("   示例：gong,你好")

    # 速度自动循环：每个速度6次，共18次
    speed_list = ["slow", "medium", "high"]
    repeat_per_speed = 6

    # 确认信息
    print("\n--- 请确认以下信息 ---")
    print(f"采集者（worker）：{worker}")
    print(f"中文标签（label）：{label}")
    print(f"速度计划：slow×{repeat_per_speed} → medium×{repeat_per_speed} → high×{repeat_per_speed}，共18次")
    print(f"起始动作标志（sample_id）：gt_{sample_counter:03d}")
    confirm = input("确认无误？(y/n)：").strip().lower()
    if confirm != 'y':
        print("已取消，请重新运行脚本。")
        return

    # 初始化速度调度
    speed_index = 0
    speed_count = 0
    
    total_rounds = len(speed_list) * repeat_per_speed  # 18
    current_round = 0

    print("\n开始数据采集。输入数据后按回车提交，输入 'q' 退出。")
    print("数据格式：timestamp_ms,hand_type,f_0_s1_acc_x,...dorsal_mag_z")
    print("多条数据格式：[DATA]数据1[DATA]数据2...\n")

    while current_round < total_rounds:
        # 计算当前速度
        speed = speed_list[speed_index]
        remaining = repeat_per_speed - speed_count
        print(f"\n【第 {current_round+1}/{total_rounds} 次】当前速度：{speed}（该速度还剩 {remaining} 次）")
        print(">>> 请输入数据：")
        lines = []
        while True:
            try:
                line = input()
                if line.strip().lower() == 'q':
                    wb.save(EXCEL_PATH)
                    print(f"数据已保存到 {EXCEL_PATH}，程序退出。")
                    return
                lines.append(line)
            except EOFError:
                break

        input_str = "\n".join(lines)
        if not input_str.strip():
            continue

        # 解析数据
        data_rows = parse_input_data(input_str)
        if not data_rows:
            print("未解析到有效数据。")
            continue

        # 生成当前批次的动作标志
        current_sample_id = f"gt_{sample_counter:03d}"

        # 获取写入起始行
        start_row = get_target_start_row(ws)
        for idx, data_parts in enumerate(data_rows):
            row_num = start_row + idx
            row_data = [worker, speed, label, current_sample_id] + data_parts

            # 从第2列（B列）开始写入，跳过第1列（A列）
            for col_num, value in enumerate(row_data, 2):
                if col_num > total_cols:
                    break
                try:
                    if "." in value:
                        ws.cell(row=row_num, column=col_num, value=float(value))
                    else:
                        ws.cell(row=row_num, column=col_num, value=int(value))
                except (ValueError, TypeError):
                    ws.cell(row=row_num, column=col_num, value=value)

        # 保存
        try:
            wb.save(EXCEL_PATH)
            print(f"✅ 成功写入 {len(data_rows)} 行数据！")
            print(f"   本次动作标志：{current_sample_id}，速度：{speed}，写入起始行：{start_row}")
            sample_counter += 1
            current_round += 1
            speed_count += 1
            # 当前速度录满，切换到下一个速度
            if speed_count >= repeat_per_speed:
                speed_index += 1
                speed_count = 0
        except Exception as e:
            print(f"❌ 保存Excel文件出错：{e}")
            print("请确保文件未被其他程序打开。")

    # 18次全部完成
    print(f"\n🎉 该动作「{label}」的18次采集已全部完成！")

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("错误：未安装openpyxl库。请运行 'pip install openpyxl' 安装。")
        sys.exit(1)
    
    main()